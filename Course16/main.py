import csv
import openpyxl
import json
import requests
import xml.etree.ElementTree as Et
import PyPDF2
from bs4 import BeautifulSoup

# Working with a CSV file
print(csv.list_dialects())
csv_rows = []

with open("data.csv") as csv_file:
    reader = csv.reader(csv_file, delimiter=";", quoting=csv.QUOTE_NONE)
    for row in reader:
        csv_rows.append(row)
        price = row[1]
        print(price)

with open("data.csv", "a") as csv_file:
    writer = csv.writer(csv_file, delimiter=";")
    row = ["iPad", "900", "2"]
    writer.writerow(row)

# Working with an Excel file
wb = openpyxl.Workbook()
ws = wb.active
for row in csv_rows:
    ws.append(row)

wb.save("file.xlsx")

wb.close()

wb = openpyxl.load_workbook("file.xlsx")
ws = wb.active
print(ws["A2"].value)
wb.close()

# Working with a JSON file
with open("example.json", "r") as file:
    json_file = json.load(file)
    q1 = json_file["quiz"]["sport"]["q1"]
    question = q1["question"]
    print(question)
    answer = q1["options"][2]
    print(answer)

with open("result.json", "w") as file:
    object_to_be_written = {"header": csv_rows[0]}
    json.dump(object_to_be_written, file)

# Working with an XML file
xml = Et.parse("file.xml")
it = xml.getroot().findall("country")
for i in it:
    print(i.get(key="name"))
    print(i.find("gdppc").text)

# Working with an HTML file
response = requests.get("https://google.com")
soup = BeautifulSoup(response.content, features="html.parser")
link_list = soup.find_all("a")
print(link_list)

# Working with a PDF file
with open("file.pdf", "rb") as pdf_file:
    reader = PyPDF2.PdfReader(pdf_file)
    for page in reader.pages:
        print(page.extract_text())




